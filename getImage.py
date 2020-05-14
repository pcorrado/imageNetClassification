
import openpyxl
import requests

file = '/Users/philipcorrado/Box Sync/Code/imageNetClassification/links.xlsx'
wb = openpyxl.load_workbook(file)
ws = wb['Sheet1']
for row in range(1,41):
    print(ws.cell(row=row, column=1).value)  # This will fail if there is no
    print(ws.cell(row=row, column=1).hyperlink.target)  # This will fail if there is no

imageURL='https://farm1.static.flickr.com/171/370630966_5fe1b748ac.jpg'
r = requests.get(imageURL, allow_redirects=True)
open('test.jpeg', 'wb').write(r.content)
